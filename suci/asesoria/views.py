from django.shortcuts import render, HttpResponse, redirect
from django.http import FileResponse, Http404, JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 
from io import BytesIO
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from .models import Denuncia, RegistroFilmico
from .forms import FormDenuncia, FormRegistroF
from django.utils import timezone  # Agrega esta línea

## VIEWS DE DENUNCIAS
def denuncia(request, accion):

    denuncias = Denuncia.objects.all()
    form = FormDenuncia()
    
    #   CONSULTA POR TODOS 
    if accion == 'todos':
        form = FormDenuncia()
        
    #   AGREGAR DENUNCIA
    if accion == 'agregar':
        if request.method =='POST':
            form = FormDenuncia(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/asesoria/denuncias/todos')
                
    if accion == 'editar':
        if request.method == 'POST':
            form = FormDenuncia(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                ente_m = form.cleaned_data['ente']
                nombres_d_m = form.cleaned_data['nombres_d']
                apellidos_d_m = form.cleaned_data['apellidos_d']
                cedula_d_m = form.cleaned_data['cedula_d']
                telefono_m = form.cleaned_data['telefono']
                email_m = form.cleaned_data['email']
                direccion_d_m = form.cleaned_data['direccion_d']
                nombres_denunciado_m = form.cleaned_data['nombres_denunciado']
                apellidos_denunciado_m = form.cleaned_data['apellidos_denunciado']
                cedula_denunciado_m = form.cleaned_data['cedula_denunciado']
                motivo_m = form.cleaned_data['motivo']
                zona_m = form.cleaned_data['zona']
                fecha_denuncia_m = form.cleaned_data['fecha_denuncia']

                if estatus_m ==  "ACTIVA":
                    fecha_denuncia_m = timezone.now()
                else:
                    fecha_denuncia_m = None
                    
                fecha_incidente_m = form.cleaned_data['fecha_incidente']
                
                try:
                    denuncia = Denuncia.objects.get(id=id_m)
                except Denuncia.DoesNotExist:
                    raise ValueError("Denuncia no existe.")
                    
                denuncia.estatus = estatus_m
                denuncia.ente = ente_m
                denuncia.nombres_d = nombres_d_m
                denuncia.apellidos_d = apellidos_d_m
                denuncia.cedula_d = cedula_d_m
                denuncia.telefono = telefono_m
                denuncia.email = email_m
                denuncia.direccion_d = direccion_d_m
                denuncia.nombres_denunciado = nombres_denunciado_m
                denuncia.apellidos_denunciado = apellidos_denunciado_m
                denuncia.cedula_denunciado = cedula_denunciado_m
                denuncia.motivo = motivo_m
                denuncia.zona = zona_m
                denuncia.fecha_denuncia = fecha_denuncia_m
                denuncia.fecha_incidente = fecha_incidente_m
                denuncia.save()
        else:
             
             form = FormDenuncia()

        #   CONSULTAR DENUNCIA POR CEDULA
    if accion == 'consultar':
        if 'cedula' in request.GET:
            cedula = request.GET['cedula']
            denuncias = Denuncia.objects.filter(cedula_d=cedula)
            form = FormDenuncia()
    
    
    context = {
        "denuncias": denuncias, 
        "form": form
    }
    return render(request, "denuncias.html", context)

def obtener_datos_denuncia_edicion(request, denuncia_id):
    try:
        denuncias = Denuncia.objects.get(id=denuncia_id)
        data = {
            'id': denuncias.id,
            'estatus': denuncias.estatus,
            'ente': denuncias.ente,
            'nombres_d': denuncias.nombres_d,
            'apellidos_d': denuncias.apellidos_d,
            'cedula_d': denuncias.cedula_d,
            'telefono': denuncias.telefono,
            'email': denuncias.email,
            'direccion_d': denuncias.direccion_d,
            'nombres_denunciado': denuncias.nombres_denunciado,
            'apellidos_denunciado': denuncias.apellidos_denunciado,
            'cedula_denunciado': denuncias.cedula_denunciado,
            'motivo': denuncias.motivo,
            'zona': denuncias.zona,
            'fecha_denuncia': denuncias.fecha_denuncia,
            'fecha_incidente': denuncias.fecha_incidente,
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Denuncia.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  
    
## DESCARGA DE DENUNCIAS EXCEL

def descargar_excel_denuncias(request):
    # Filtra los datos del modelo para generar el archivo Excel
    denuncias = Denuncia.objects.all()

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:P1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Denuncias del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    
    ws.append(["Estatus", "Ente", "Nombres del denunciante", "Apellidos del denunciante", "Cedula del denunciante", "Telefono", "Email", "Direccion", "Nombres del denunciado", "Apellidos del denunciado", "Cedula del denunciado", "Motivo", "Zona", "Fecha de la denuncia", "Fecha del incidente"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 10
    columnas['B'].width = 15
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 25
    columnas['F'].width = 25
    columnas['G'].width = 25
    columnas['H'].width = 25
    columnas['I'].width = 25
    columnas['J'].width = 25
    columnas['K'].width = 25
    columnas['L'].width = 25
    columnas['M'].width = 25
    columnas['N'].width = 25
    columnas['O'].width = 20
    columnas['P'].width = 20
    
    for dato in denuncias:
       
       ws.append([dato.estatus, dato.ente, dato.nombres_d, dato.apellidos_d, dato.cedula_d, dato.telefono, dato.email, dato.direccion_d, dato.nombres_denunciado, dato.apellidos_denunciado, dato.cedula_denunciado, dato.motivo, dato.zona, dato.fecha_denuncia, dato.fecha_incidente])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="Denuncias.xlsx")


## VIEWS DE REGISTRO FILMICO 
def registroFilmico(request, accion):

    registroF = RegistroFilmico.objects.all()
    
    #   CONSULTA POR TODOS 
    if accion == 'todos':
        form = FormRegistroF()
        
    #   AGREGAR REGISTRO
    if accion == 'agregar':
        if request.method =='POST':
            form = FormRegistroF(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/asesoria/registros/todos')
            
    if accion == 'editar':
        if request.method == 'POST':
            form = FormRegistroF(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                direccion_m = form.cleaned_data['direccion']
                camara_m = form.cleaned_data['camara']
                motivo_solicitud_m = form.cleaned_data['motivo_solicitud']
                ente_solicita_m = form.cleaned_data['ente_solicita']
                fecha_solicitud_m = form.cleaned_data['fecha_solicitud']
                fecha_culminacion_m = form.cleaned_data['fecha_culminacion']
                try:
                    registrosf = RegistroFilmico.objects.get(id=id_m)
                except RegistroFilmico.DoesNotExist:
                    raise ValueError("Registro no existe.")
                    
                registrosf.estatus = estatus_m
                registrosf.direccion = direccion_m
                registrosf.camara = camara_m
                registrosf.motivo_solicitud = motivo_solicitud_m
                registrosf.ente_solicita = ente_solicita_m
                registrosf.fecha_solicitud = fecha_solicitud_m
                registrosf.fecha_culminacion = fecha_culminacion_m
                registrosf.save()
        else:
             
             form = FormRegistroF()

        #   CONSULTAR POR CAMARA
    if accion == 'consultar':
        if 'camara' in request.GET:
            camara = request.GET['camara']
            registroF = RegistroFilmico.objects.filter(camara=camara)
            form = FormRegistroF()
    
    
    context = {
        "registroF": registroF, 
        "form": form
    }
    return render(request, "registroFilmico.html", context)

def obtener_datos_registro_edicion(request, registro_id):
    try:
        registroF = RegistroFilmico.objects.get(id=registro_id)
        data = {
            'id': registroF.id,
            'estatus': registroF.estatus,
            'direccion': registroF.direccion,
            'camara': registroF.camara,
            'motivo_solicitud': registroF.motivo_solicitud,
            'ente_solicita': registroF.ente_solicita,
            'fecha_solicitud': registroF.fecha_solicitud,
            'fecha_culminacion': registroF.fecha_culminacion,

            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except RegistroFilmico.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  

## DESCARGA DE EXCEL REGISTRO
    
def descargar_excel_registros(request):
    # Filtra los datos del modelo para generar el archivo Excel
    registro = RegistroFilmico.objects.all()

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:G1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Registros Filmicos del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    
    ws.append(["Estatus", "Direccion", "Camara", "Motivo de solicitud", "Ente que solicita", "Fecha de la solicitud", "Fecha de culminacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 30
    columnas['F'].width = 30
    columnas['G'].width = 30
    
    for dato in registro:
       
       ws.append([dato.estatus, dato.direccion, dato.camara, dato.motivo_solicitud, dato.ente_solicita, dato.fecha_solicitud, dato.fecha_culminacion])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="RegistroFilmico.xlsx")
