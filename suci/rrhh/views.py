from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, HttpResponse, redirect
from django.http import FileResponse, Http404, JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 
from io import BytesIO
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from .forms import GeneroFormSexo, FormEstadoC, FormTallaC, FormTallaP, FormTallaZ, FormGradoI, FormTipoPersonal, FormCargo, FormDepartamento, FormSedes, FormPersonal, FormMovimientoPersonal, FormEditarPersonal, FormBienes, FormBonoGuerra, FormCestaTicket, FormSueldoMinimo, FormPrima, FormPrimaAdd
from datetime import datetime
from django.utils import timezone  # Agrega esta línea
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import datetime

# Create your views here.

from rrhh.models import Personal, Cargo, Sexo, EstadoCivil, Sangre, TallasCamisa, TallasPantalon, TallasZapatos, Grado, TipoPersonal, Departamento, Sedes, Nacionalidad, Bienes, Sueldos, Tasa, PrimaProfesionalismo, SueldoMinimo, CestaTicket, BonoGuerra

# Create your views here.

def obtenerPrimaProfesionalismo(grado_instruccion, sueldo_minimo):
    prima = PrimaProfesionalismo.objects.get(titulo=grado_instruccion)
    porcentaje = prima.porcentaje
    monto = (sueldo_minimo * porcentaje) / 100
    return monto

def personal(request, accion):

    personales = Personal.objects.all()
    
    #   CONSULTA POR TODOS 
    if accion == 'todos':
        form = FormPersonal()
        
    #   AGREGAR PERSONAL
    if accion == 'agregar':
        if request.method =='POST':
            form = FormPersonal(request.POST)
            if form.is_valid():
                grado_instruccion = form.cleaned_data['grado_instruccion']
                form.save()
                id = form.instance.id
                sueldos = SueldoMinimo.objects.first()
                sueldo_minimo = sueldos.monto
                prima_profesionalismo = obtenerPrimaProfesionalismo(grado_instruccion, sueldo_minimo)
                cesta_ticket = CestaTicket.objects.get(id=1)
                bono_guerra = BonoGuerra.objects.get(id=1)
                id_personal = Personal.objects.get(id=id)
                
                formsueldo = Sueldos(
                    estatus = 'ACTIVO',
                    sueldo_base = sueldo_minimo,
                    prima_profesionalismo = prima_profesionalismo,
                    p_discapacidad = 0,
                    p_hijos_menor_12 = 0,
                    p_hijas_menor_12 = 0,
                    p_hijos_12_18 = 0, 
                    p_hijos_discapacidad = 0,
                    n_fasmij = 0,
                    cesta_t = 1,
                    monto_t = cesta_ticket,
                    b_guerra = 1,
                    monto_b = bono_guerra,
                    personal = id_personal
                )
                formsueldo.save()
                return redirect('/rrhh/personal/todos')
            
    #   MOVIMIENTO DE PERSONAL
    if accion == 'movimiento':
        if request.method == 'POST':
            form = FormMovimientoPersonal(request.POST)
            if form.is_valid():
                estatus_m = form.cleaned_data['estatus']
                nombre_m = form.cleaned_data['nombres']
                apellido_m = form.cleaned_data['apellidos']
                nacionalidad_m = form.cleaned_data['nacionalidad']
                cedula_m = form.cleaned_data['cedula']
                departamento_m = form.cleaned_data['departamento']
                sede_m = form.cleaned_data['sede']
                motivo_m = form.cleaned_data['motivo']
                if estatus_m ==  "RETIRADO":
                    fecha_retiro_m = timezone.now()
                else:
                    fecha_retiro_m = None
                
                try:
                    personal = Personal.objects.get(cedula=cedula_m)
                except Personal.DoesNotExist:
                    raise ValueError("El personal no existe.")
                
                personal.estatus = estatus_m
                personal.nombres = nombre_m
                personal.apellidos = apellido_m
                personal.nacionalidad = nacionalidad_m
                personal.cedula = cedula_m
                personal.departamento = departamento_m
                personal.sede = sede_m
                personal.motivo = motivo_m
                personal.fecha_retiro = fecha_retiro_m
                personal.save()
                
                #return redirect('lista_personal')
        else:
            form = FormMovimientoPersonal()
        

    #   CONSULTAR PERSONAL POR CEDULA
    if accion == 'consultar':
        if 'cedula' in request.GET:
            cedula = request.GET['cedula']
            personales = Personal.objects.filter(cedula=cedula)
            form = FormPersonal()
    

    #   EDICION DE PERSONAL
    if accion == 'editar':
        if request.method == 'POST':
            form = FormEditarPersonal(request.POST)
            if form.is_valid():
                estatus_m = form.cleaned_data['estatus']
                nombre_m = form.cleaned_data['nombres']
                apellido_m = form.cleaned_data['apellidos']
                nacionalidad_m = form.cleaned_data['nacionalidad']
                cedula_m = form.cleaned_data['cedula']
                sexo_m = form.cleaned_data['sexo']
                fecha_nac_m = form.cleaned_data['fecha_nac']
                edad_m = form.cleaned_data['edad']
                telefono_m = form.cleaned_data['telefono']
                estado_civil_m = form.cleaned_data['estado_civil']
                conyugue_m = form.cleaned_data['conyugue']
                cedula_conyugue_m = form.cleaned_data['cedula_conyugue']
                sangre_m = form.cleaned_data['tipo_sangre']
                discapacitado_m = form.cleaned_data['discapacitado']
                talla_camisa_m = form.cleaned_data['talla_camisa']
                talla_pantalon_m = form.cleaned_data['talla_pantalon']
                talla_zapato_m = form.cleaned_data['talla_zapato']
                direccion_m = form.cleaned_data['direccion']
                nro_cuenta_m = form.cleaned_data['nro_cuenta']
                email_m = form.cleaned_data['email']
                grado_instruccion_m = form.cleaned_data['grado_instruccion']
                estudias_m = form.cleaned_data['estudias']
                comision_servicio_m = form.cleaned_data['comision_servicio']
                pnb_m = form.cleaned_data['pnb']
                tipo_personal_m = form.cleaned_data['tipo_personal']
                cargo_m = form.cleaned_data['cargo']
                fecha_ingreso_911_m = form.cleaned_data['fecha_ingreso_911']
                fecha_ingreso_apn_m = form.cleaned_data['fecha_ingreso_apn']
                contratos_m = form.cleaned_data['contratos']
                departamento_m = form.cleaned_data['departamento']
                nino_menor_12_m = form.cleaned_data['nino_menor_12']
                edades1_m = form.cleaned_data['edades1']
                hijos_13_18_m = form.cleaned_data['hijos_13_18']
                edades2_m = form.cleaned_data['edades2']
                nina_menor_12_m = form.cleaned_data['nina_menor_12']
                edades3_m = form.cleaned_data['edades3']
                hijos_discapacidad_m = form.cleaned_data['hijos_discapacidad']
                edades4_m = form.cleaned_data['edades4']
                sede_m = form.cleaned_data['sede']
                motivo_m = form.cleaned_data['motivo']
                fasmij_m = form.cleaned_data['fasmij']
                parentezco1_m = form.cleaned_data['parentezco1']
                beneficiario1_m = form.cleaned_data['beneficiario1']
                cedula1_m = form.cleaned_data['cedula1']
                direccion1_m = form.cleaned_data['direccion1']
                parentezco2_m = form.cleaned_data['parentezco2']
                beneficiario2_m = form.cleaned_data['beneficiario2']
                cedula2_m = form.cleaned_data['cedula2']
                direccion2_m = form.cleaned_data['direccion2']
                parentezco3_m = form.cleaned_data['parentezco3']
                beneficiario3_m = form.cleaned_data['beneficiario3']
                cedula3_m = form.cleaned_data['cedula3']
                direccion3_m = form.cleaned_data['direccion3']
                #personal = Personal.objects.get(cedula=cedula_m)
                try:
                    personal = Personal.objects.get(cedula=cedula_m)
                except Personal.DoesNotExist:
                    raise ValueError("El personal no existe.")
                
                personal.estatus = estatus_m
                personal.nombres = nombre_m
                personal.apellidos = apellido_m
                personal.nacionalidad = nacionalidad_m
                personal.cedula = cedula_m
                personal.sexo = sexo_m
                personal.fecha_nac = fecha_nac_m
                personal.edad = edad_m
                personal.telefono = telefono_m
                personal.estado_civil = estado_civil_m
                personal.conyugue = conyugue_m
                personal.cedula_conyugue = cedula_conyugue_m
                personal.tipo_sangre = sangre_m
                personal.discapacitado = discapacitado_m
                personal.talla_camisa = talla_camisa_m
                personal.talla_pantalon = talla_pantalon_m
                personal.talla_zapato = talla_zapato_m
                personal.direccion = direccion_m
                personal.nro_cuenta = nro_cuenta_m
                personal.email = email_m
                personal.grado_instruccion = grado_instruccion_m
                personal.estudias = estudias_m
                personal.comision_servicio = comision_servicio_m
                personal.pnb = pnb_m
                personal.tipo_personal = tipo_personal_m
                personal.cargo = cargo_m
                personal.fecha_ingreso_911 = fecha_ingreso_911_m
                personal.fecha_ingreso_apn = fecha_ingreso_apn_m
                personal.contratos = contratos_m
                personal.departamento = departamento_m
                personal.nino_menor_12 = nino_menor_12_m
                personal.edades1 = edades1_m
                personal.hijos_13_18 = hijos_13_18_m
                personal.edades2 = edades2_m
                personal.nina_menor_12 = nina_menor_12_m
                personal.edades3 = edades3_m
                personal.hijos_discapacidad = hijos_discapacidad_m
                personal.edades4 = edades4_m
                personal.sede = sede_m
                personal.motivo = motivo_m
                personal.fasmij = fasmij_m
                personal.parentezco1 = parentezco1_m
                personal.beneficiario1 = beneficiario1_m
                personal.cedula1 = cedula1_m
                personal.direccion1 = direccion1_m
                personal.parentezco2 = parentezco2_m
                personal.beneficiario2 = beneficiario2_m
                personal.cedula2 = cedula2_m
                personal.direccion2 = direccion2_m
                personal.parentezco3 = parentezco3_m
                personal.beneficiario3 = beneficiario3_m
                personal.cedula3 = cedula3_m
                personal.direccion3 = direccion3_m
                personal.save()
                #return redirect('lista_personal')
        else:
            form = FormEditarPersonal()
        
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    context = {
        "personales": personales, 
        "cargos": cargos, 
        "sexos": sexos, 
        "estados": estados, 
        "sangres": sangres, 
        "talla_camisa": talla_camisa, 
        "talla_pantalon": talla_pantalon, 
        "talla_zapato": talla_zapato, 
        "grado_instruccion": grado_instruccion, 
        "tipo_personal": tipo_personal, 
        "departamentos": departamentos, 
        "sedes": sedes, 
        "nacionalidades": nacionalidades, 
        "form": form
    }
    return render(request, "ingresoPersonal.html", context)


#   VISTA PARA ACTUALIZAR LOS MOVIMIENTO DEL PERSONAL
def movimientoPersonal(request):
    if request.method == 'POST':
        form = FormMovimientoPersonal(request.POST)
        if form.is_valid():
            estatus_m = form.cleaned_data['estatus']
            nombre_m = form.cleaned_data['nombres']
            apellido_m = form.cleaned_data['apellidos']
            nacionalidad_m = form.cleaned_data['nacionalidad']
            cedula_m = form.cleaned_data['cedula']
            departamento_m = form.cleaned_data['departamento']
            sede_m = form.cleaned_data['sede']
            motivo_m = form.cleaned_data['motivo']
            #personal = Personal.objects.get(cedula=cedula_m)
            try:
                personal = Personal.objects.get(cedula=cedula_m)
            except Personal.DoesNotExist:
                raise ValueError("El personal no existe.")
            
            personal.estatus = estatus_m
            personal.nombres = nombre_m
            personal.apellidos = apellido_m
            personal.nacionalidad = nacionalidad_m
            personal.cedula = cedula_m
            personal.departamento = departamento_m
            personal.sede = sede_m
            personal.motivo = motivo_m
            personal.save()
            #return redirect('lista_personal')
    else:
        form = FormMovimientoPersonal()
    personales = Personal.objects.all()
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "ingresoPersonal.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades, "form": form})


def consultaPersonal(request, cedula):
    
    personales = Personal.objects.filter(cedula=cedula)
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "ingresoPersonal.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades})


def obtener_datos(request, personal_id):
    try:
        personal = Personal.objects.get(id=personal_id)
        data = {
            'id': personal.id,
            'estatus': personal.estatus,
            'nombres': personal.nombres,
            'apellidos': personal.apellidos,
            'cedula': personal.cedula,
            'nacionalidad': personal.nacionalidad.id,
            'departamento': personal.departamento.id,
            'sede': personal.sede.id,
            'motivo': personal.motivo,
            'fecha_retiro': personal.fecha_retiro,
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    
    
def obtener_datos_edicion(request, personal_id):
    try:
        personal = Personal.objects.get(id=personal_id)
        data = {
            'id': personal.id,
            'estatus': personal.estatus,
            'nombres': personal.nombres,
            'apellidos': personal.apellidos,
            'cedula': personal.cedula,
            'nacionalidad': personal.nacionalidad.id,
            'sexo': personal.sexo.id,
            'fecha_nac': personal.fecha_nac,
            'edad': personal.edad,
            'telefono': personal.telefono,
            'estado_civil': personal.estado_civil.id,
            'conyugue': personal.conyugue,
            'ciconyugue': personal.cedula_conyugue,
            'tipo_sangre': personal.tipo_sangre.id,
            'discapacitado': personal.discapacitado,
            'talla_camisa': personal.talla_camisa.id,
            'talla_pantalon': personal.talla_pantalon.id,
            'talla_zapato': personal.talla_zapato.id,
            'direccion': personal.direccion,
            'nro_cuenta': personal.nro_cuenta,
            'email': personal.email,
            'grado_instruccion': personal.grado_instruccion.id,
            'estudias' : personal.estudias,
            'comision_servicio' : personal.comision_servicio,
            'pnb': personal.pnb,
            'tipo_personal': personal.tipo_personal.id,
            'cargo': personal.cargo.id,
            'fecha_ingreso_911': personal.fecha_ingreso_911,
            'fecha_ingreso_apn': personal.fecha_ingreso_apn,
            'contratos': personal.contratos,
            'departamento': personal.departamento.id,
            'nino_menor_12': personal.nino_menor_12,
            'edades1': personal.edades1,
            'hijos_13_18': personal.hijos_13_18,
            'edades2': personal.edades2,
            'nina_menor_12': personal.nina_menor_12,
            'edades3': personal.edades3,
            'hijos_discapacidad': personal.hijos_discapacidad,
            'edades4': personal.edades4,
            'sede': personal.sede.id,
            'motivo': personal.motivo,
            'fasmij': personal.fasmij,
            'parentezco1': personal.parentezco1,
            'parentezco2': personal.parentezco2,
            'parentezco3': personal.parentezco3,
            'beneficiario1': personal.beneficiario1,
            'beneficiario2': personal.beneficiario2,
            'beneficiario3': personal.beneficiario3,
            'cedula1': personal.cedula1,
            'cedula2': personal.cedula2,
            'cedula3': personal.cedula3,
            'direccion1': personal.direccion1,
            'direccion2': personal.direccion2,
            'direccion3': personal.direccion3,
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    
def descargar_excel(request):
    # Filtra los datos del modelo para generar el archivo Excel
    personales = Personal.objects.all()

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:N1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Personal Completo del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título

    ws.append(["Estatus", "Nombres", "Apellidos", "Nacionalidad", "Cedula", "Genero", "Fecha Nac", "Edad", "Telefono", "Estado Civil", "Conyugue", "CI Conyugue", "Tipo Sangre", "Discapacitado", "Camisa", "Pantalon", "Zapatos", "Direccion", "Nro Cuenta", "Email", "Grado Instruccion", "Estudias", "Comision", "Pnb", "Tipo", "Cargo", "Fecha Ingreso 911", "Fecha Ingreso APN", "Contratos", "Departamentos", "Niños < 12", "Edad", "Hijos > 13", "Edad", "Niña < 12", "Edad", "Hijos Discapacidad", "Edad", "Motivo Retiro", "Fecha Retiro", "Sede", "Fasmij", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Fecha Creacion", "Fecha Actualizacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    columnas['F'].width = 15
    columnas['G'].width = 12
    columnas['H'].width = 5
    columnas['I'].width = 12
    columnas['J'].width = 12
    columnas['K'].width = 15
    columnas['L'].width = 10
    columnas['M'].width = 12
    columnas['N'].width = 12
    columnas['O'].width = 10
    columnas['P'].width = 10
    columnas['Q'].width = 10
    columnas['R'].width = 30
    columnas['S'].width = 25
    columnas['T'].width = 25
    columnas['U'].width = 17
    columnas['V'].width = 10
    columnas['W'].width = 10
    columnas['X'].width = 10
    columnas['Y'].width = 10
    columnas['Z'].width = 20
    columnas['AA'].width = 18
    columnas['AB'].width = 18
    columnas['AC'].width = 10
    columnas['AD'].width = 25
    columnas['AE'].width = 18
    columnas['AF'].width = 8
    columnas['AG'].width = 18
    columnas['AH'].width = 8
    columnas['AI'].width = 18
    columnas['AJ'].width = 8
    columnas['AK'].width = 18
    columnas['AL'].width = 8
    columnas['AM'].width = 25
    columnas['AN'].width = 15
    columnas['AO'].width = 15
    columnas['AP'].width = 8
    columnas['AQ'].width = 15
    columnas['AR'].width = 20
    columnas['AS'].width = 15
    columnas['AT'].width = 30
    columnas['AU'].width = 15
    columnas['AV'].width = 20
    columnas['AW'].width = 30
    columnas['AX'].width = 30
    columnas['AY'].width = 15
    columnas['AZ'].width = 20
    columnas['BA'].width = 15
    columnas['BB'].width = 30
    columnas['BC'].width = 20
    columnas['BD'].width = 20
    for dato in personales:
       discapacitado =  dato.discapacitado
       if (discapacitado ==  True):
           discapacitado = 'SI'
       else:
           discapacitado = 'NO'
           
       estudias =  dato.estudias
       if (estudias ==  True):
           estudias = 'SI'
       else:
           estudias = 'NO' 
           
       comision =  dato.comision_servicio
       if (comision ==  True):
           comision = 'SI'
       else:
           comision = 'NO' 
           
       pnb =  dato.pnb
       if (pnb ==  True):
           pnb = 'SI'
       else:
           pnb = 'NO' 
           
       fasmij =  dato.fasmij
       if (fasmij ==  True):
           fasmij = 'SI'
       else:
           fasmij = 'NO' 
           
       ws.append([dato.estatus, dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.sexo.sexo, dato.fecha_nac, dato.edad, dato.telefono, dato.estado_civil.estado_civil, dato.conyugue, dato.cedula_conyugue, dato.tipo_sangre.tipo_sangre, discapacitado, dato.talla_camisa.talla_camisa, dato.talla_pantalon.talla_pantalon, dato.talla_zapato.talla_zapato, dato.direccion, dato.nro_cuenta, dato.email, dato.grado_instruccion.grado_instruccion, estudias, comision, pnb, dato.tipo_personal.tipo_personal, dato.cargo.cargo, dato.fecha_ingreso_911, dato.fecha_ingreso_apn, dato.contratos, dato.departamento.departamento, dato.nino_menor_12, dato.edades1, dato.hijos_13_18, dato.edades2, dato.nina_menor_12, dato.edades3, dato.hijos_discapacidad, dato.edades4, dato.motivo, dato.fecha_retiro, dato.sede.sede, fasmij, dato.parentezco1, dato.beneficiario1, dato.cedula1, dato.direccion1, dato.parentezco2, dato.beneficiario2, dato.cedula2, dato.direccion2, dato.parentezco3, dato.beneficiario3, dato.cedula3, dato.direccion3])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="datosPersonal.xlsx")


    # GENERAR PDF 

def generar_pdf(request):
    # Consulta al modelo para obtener todos los datos
    datos = Personal.objects.all()

    # Crear el objeto HttpResponse con el tipo de contenido correcto
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Personal.pdf"'

    # Crear el objeto PDF
    pdf = SimpleDocTemplate(response, pagesize=letter, topMargin=20)
    elements = []
    
    
    # Agregar título centrado, en negrita y tamaño de letra 12
    title = "Reporte de Datos del Personal"
    title_style = getSampleStyleSheet()["Title"]
    title_style.alignment = 1  # 0=Left, 1=Center, 2=Right
    title_style.fontSize = 14
    title_text = Paragraph(title, title_style)
    elements.append(Spacer(1, 6))
    elements.append(title_text)
    
    # Agregar espacio en blanco entre el título y la tabla
    elements.append(Spacer(1, 2))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    # Agregar fecha debajo del título sin negrita y tamaño de letra 10
    now = datetime.datetime.now()
    fecha = "Fecha: %s / Hora: %s" % (now.strftime("%Y-%m-%d"), now.strftime("%H:%M"))
    fecha_style = getSampleStyleSheet()["Normal"]
    fecha_style.alignment = 1
    fecha_style.fontSize = 8
    fecha_text = Paragraph(fecha, fecha_style)
    elements.append(fecha_text)
    
    #Agregar espacio en blanco entre la fecha y la tabla
    elements.append(Spacer(1, 15))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    

    # Crear los datos para la tabla
    data = [['Nombres', 'Apellidos', 'Nacionalidad', 'Cédula', 'Teléfono', 'Cargo', 'Departamentos']]  # Encabezados de la tabla
    for dato in datos:

        discapacitado =  dato.discapacitado
        if (discapacitado ==  True):
           discapacitado = 'SI'
        else:
           discapacitado = 'NO'
           
        estudias =  dato.estudias
        if (estudias ==  True):
           estudias = 'SI'
        else:
           estudias = 'NO' 
           
        comision =  dato.comision_servicio
        if (comision ==  True):
           comision = 'SI'
        else:
           comision = 'NO' 
           
        pnb =  dato.pnb
        if (pnb ==  True):
           pnb = 'SI'
        else:
           pnb = 'NO' 
           
        fasmij =  dato.fasmij
        if (fasmij ==  True):
           fasmij = 'SI'
        else:
           fasmij = 'NO' 
           
        data.append([dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.telefono, dato.cargo.cargo])  # Agregar datos del modelo

    # Crear la tabla y aplicar estilos
    table = Table(data)
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    table.setStyle(style)

    # Agregar la tabla al PDF
    elements.append(table)

    # Construir el PDF y devolver la respuesta
    pdf.build(elements)
    return response
    
    
 #VISTA DE CARGO 
    
def cargo():
    cargos = Cargo.objects.all()
    return cargos

#VISTA DE SEXO 
def sexo():
    sexos = Sexo.objects.all()
    return sexos

def estado():
    estados = EstadoCivil.objects.all()
    return estados
    
    #return render(request, "estadoCivil.html", {"estados": estados})

def sangre():
    sangres = Sangre.objects.all()
    return sangres


    
def tallaCamisa():
    talla_camisa = TallasCamisa.objects.all()
    return talla_camisa
    


def tallaPantalon():
    talla_pantalon = TallasPantalon.objects.all()
    return talla_pantalon


    
def tallaZapatos():
    talla_zapato = TallasZapatos.objects.all()
    return talla_zapato



def gradoInstruccion():
    grado_instruccion = Grado.objects.all()
    return grado_instruccion



def tipoPersonal():
    tipo_personal = TipoPersonal.objects.all()
    return tipo_personal



def departamento():
    departamentos = Departamento.objects.all()
    return departamentos


    
def sede():
    sedes = Sedes.objects.all()
    return sedes


def nacionalidad():
    nacionalidades = Nacionalidad.objects.all()
    return nacionalidades
    

def estadoCivil(request):

    if request.method =='POST':
        
        form = FormEstadoC(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/rrhh/estadocivil')
    else: 
            form = FormEstadoC()

    estadosCiviles=EstadoCivil.objects.all()

    return render(request, "mantenimientos/estadoCivil.html", {"estadosCiviles":estadosCiviles, "form":form})


def mantSexo(request):

    if request.method =='POST':
        
        form = GeneroFormSexo(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/rrhh/genero')
    else: 
            form = GeneroFormSexo()

    genero=Sexo.objects.all()
    return render(request, "mantenimientos/sexo.html", {"genero":genero, "form":form})


def mantTallaCamisa(request, accion):

    talla_camisa=TallasCamisa.objects.all()
    
    if accion == 'todos':
        form = FormTallaC()

    #   AGREGAR TALLA 
    if accion == 'agregar':
        if request.method =='POST':
            form = FormTallaC(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/rrhh/tallaCamisa/todos')

    #EDITAR TALLA     
    if accion == 'editar':
        if request.method == 'POST':
            form = FormTallaC(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                talla_camisa_m = form.cleaned_data['talla_camisa']
                
                try:
                    tallaC = TallasCamisa.objects.get(id=id_m)
                except TallasCamisa.DoesNotExist:
                    raise ValueError("Talla no existe.")
                    
                tallaC.estatus = estatus_m
                tallaC.talla_camisa = talla_camisa_m
                tallaC.save()
        else:
             
             form = FormTallaC()
            
    
    context = {
        "talla_camisa": talla_camisa,
        "form": form
    }

    return render(request, "mantenimientos/tallasCamisa.html", context)
      

def obtener_datos_tallaC_edicion(request, tallaC_id):
    try:
        tallaC = TallasCamisa.objects.get(id=tallaC_id)
        data = {
            'id': tallaC.id,
            'estatus': tallaC.estatus,
            'talla': tallaC.talla_camisa,
            
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    


def mantTallaPantalon(request,  accion):

    talla_pantalon=TallasPantalon.objects.all()
   
    if accion == 'todos':
            form = FormTallaP()

    #   AGREGAR TALLA 
    if accion == 'agregar':
        if request.method =='POST':
            form = FormTallaP(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/rrhh/tallaPantalon/todos')

    ## EDITAR TALLA      
    if accion == 'editar':
        if request.method == 'POST':
            form = FormTallaP(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                talla_pantalon_m = form.cleaned_data['talla_pantalon']
                    
                try:
                    tallaP = TallasPantalon.objects.get(id=id_m)
                except TallasPantalon.DoesNotExist:
                    raise ValueError("Talla no existe.")
                        
                tallaP.estatus = estatus_m
                tallaP.talla_pantalon = talla_pantalon_m
                tallaP.save()
        else:
                
            form = FormTallaP()
                
        
    context = {
        "talla_pantalon": talla_pantalon,
        "form": form
    }

    return render(request, "mantenimientos/tallasPantalon.html", context)


def obtener_datos_tallaP_edicion(request, tallaP_id):
    try:
        tallaP = TallasPantalon.objects.get(id=tallaP_id)
        data = {
            'id': tallaP.id,
            'estatus': tallaP.estatus,
            'talla': tallaP.talla_pantalon,
            
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    


def mantTallaZapato(request, accion):

    talla_zapato=TallasZapatos.objects.all()
   
    if accion == 'todos':
            form = FormTallaZ()

    #   AGREGAR TALLA 
    if accion == 'agregar':
        if request.method =='POST':
            form = FormTallaZ(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/rrhh/tallaZapato/todos')

    ## EDITAR TALLA   
    if accion == 'editar':
        if request.method == 'POST':
            form = FormTallaZ(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                talla_zapato_m = form.cleaned_data['talla_zapato']
                    
                try:
                    tallaZ = TallasZapatos.objects.get(id=id_m)
                except TallasZapatos.DoesNotExist:
                    raise ValueError("Talla no existe.")
                        
                tallaZ.estatus = estatus_m
                tallaZ.talla_zapato = talla_zapato_m
                tallaZ.save()
        else:
                
            form = FormTallaZ()
        
    context = {
        "talla_zapato":talla_zapato,
        "form": form
    }

    return render(request, "mantenimientos/tallasZapato.html", context)


def obtener_datos_tallaZ_edicion(request, tallaZ_id):
    try:
        tallaZ = TallasZapatos.objects.get(id=tallaZ_id)
        data = {
            'id': tallaZ.id,
            'estatus': tallaZ.estatus,
            'talla': tallaZ.talla_zapato,
            
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  

def mantGrado(request):

    if request.method =='POST':
        
        form = FormGradoI(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/rrhh/gradoInstruccion')
    else: 
            form = FormGradoI()

    grado_instruccion=Grado.objects.all()

    return render(request, "mantenimientos/gradoInstruccion.html", {"grado_instruccion":grado_instruccion, "form": form})

def mantTipoPersonal(request):

    if request.method =='POST':
        
        form = FormTipoPersonal(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/rrhh/tipoPersonal')
    else: 
            form = FormTipoPersonal()


    tipo_personal=TipoPersonal.objects.all()

    return render(request, "mantenimientos/tipoPersonal.html", {"tipo_personal":tipo_personal, "form": form})


def mantCargo(request, accion):

    cargo = Cargo.objects.all()
    form = FormCargo()
    
    if accion == 'todos':
        form = FormCargo()

    #   AGREGAR TALLA 
    if accion == 'agregar':
        if request.method =='POST':
            form = FormCargo(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/rrhh/cargo/todos')
            
    if accion == 'editar':
        if request.method == 'POST':
            form = FormCargo(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                cargo_m = form.cleaned_data['cargo']
                
                try:
                    cargoe = Cargo.objects.get(id=id_m)
                except Cargo.DoesNotExist:
                    raise ValueError("Cargo no existe.")
                    
                cargoe.estatus = estatus_m
                cargoe.cargo = cargo_m
                cargoe.save()
        else:
             
             form = FormCargo()
            
    
    context = {
        "cargo": cargo,
        "form": form
    }

    return render(request, "mantenimientos/cargo.html", context)

def obtener_datos_cargo_edicion(request, cargoE_id):
    try:
        cargo = Cargo.objects.get(id=cargoE_id)
        data = {
            'id': cargo.id,
            'estatus': cargo.estatus,
            'cargo': cargo.cargo,
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  


def mantDepartamento(request, accion):

    departamento=Departamento.objects.all()
    form = FormDepartamento()
    
    if accion == 'todos':
        form = FormDepartamento()

    #   AGREGAR TALLA 
    if accion == 'agregar':
        if request.method =='POST':
            form = FormDepartamento(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/rrhh/departamento/todos')
            
    if accion == 'editar':
        if request.method == 'POST':
            form = FormDepartamento(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                departamento_m = form.cleaned_data['departamento']
                
                try:
                    departamentoe = Departamento.objects.get(id=id_m)
                except Departamento.DoesNotExist:
                    raise ValueError("Departamento no existe.")
                    
                departamentoe.estatus = estatus_m
                departamentoe.departamento = departamento_m
                departamentoe.save()
        else:
             
             form = FormDepartamento()
            
    
    context = {
        "departamento": departamento,
        "form": form
    }

    return render(request, "mantenimientos/departamento.html", context)

def obtener_datos_departamento_edicion(request, departamento_id):
    try:
        departamento = Departamento.objects.get(id=departamento_id)
        data = {
            'id': departamento.id,
            'estatus': departamento.estatus,
            'departamento': departamento.departamento,
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  

def mantSedes(request, accion):

    sedes=Sedes.objects.all()
    form = FormSedes()
    
    if accion == 'todos':
        form = FormSedes()

    #   AGREGAR TALLA 
    if accion == 'agregar':
        if request.method =='POST':
            form = FormSedes(request.POST)
            if form.is_valid():
                form.save()
                return redirect('/rrhh/sedes/todos')
            
    if accion == 'editar':
        if request.method == 'POST':
            form = FormSedes(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                sede_m = form.cleaned_data['sede']
                
                try:
                    sede = Sedes.objects.get(id=id_m)
                except Sedes.DoesNotExist:
                    raise ValueError("Sede no existe.")
                    
                sede.estatus = estatus_m
                sede.sede = sede_m
                sede.save()
        else:
             
             form = FormSedes()
            
    
    context = {
        "sedes":sedes,
        "form": form
    }

    return render(request, "mantenimientos/sedes.html", context)

def obtener_datos_sede_edicion(request, sede_id):
    try:
        sedes=Sedes.objects.get(id=sede_id)
        data = {
            'id': sedes.id,
            'estatus': sedes.estatus,
            'sede': sedes.sede,
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  




def personalRetirado(request):
    personales = Personal.objects.filter(estatus='RETIRADO')
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "personalRetirado.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades })


# DESCARGA EXCEL PERSONAL RETIRADO

def descargar_excel_retirados(request):
    # Filtra los datos del modelo para generar el archivo Excel
    #personales = Personal.objects.all()
    personales = Personal.objects.filter(estatus='RETIRADO')

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:N1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Personal Retirado del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    
    ws.append(["Estatus", "Nombres", "Apellidos", "Nacionalidad", "Cedula", "Genero", "Fecha Nac", "Edad", "Telefono", "Estado Civil", "Conyugue", "CI Conyugue", "Tipo Sangre", "Discapacitado", "Camisa", "Pantalon", "Zapatos", "Direccion", "Nro Cuenta", "Email", "Grado Instruccion", "Estudias", "Comision", "Pnb", "Tipo", "Cargo", "Fecha Ingreso 911", "Fecha Ingreso APN", "Contratos", "Departamentos", "Niños < 12", "Edad", "Hijos > 13", "Edad", "Niña < 12", "Edad", "Hijos Discapacidad", "Edad", "Motivo Retiro", "Fecha Retiro", "Sede", "Fasmij", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Fecha Creacion", "Fecha Actualizacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    columnas['F'].width = 15
    columnas['G'].width = 12
    columnas['H'].width = 5
    columnas['I'].width = 12
    columnas['J'].width = 12
    columnas['K'].width = 15
    columnas['L'].width = 10
    columnas['M'].width = 12
    columnas['N'].width = 12
    columnas['O'].width = 10
    columnas['P'].width = 10
    columnas['Q'].width = 10
    columnas['R'].width = 30
    columnas['S'].width = 25
    columnas['T'].width = 25
    columnas['U'].width = 17
    columnas['V'].width = 10
    columnas['W'].width = 10
    columnas['X'].width = 10
    columnas['Y'].width = 10
    columnas['Z'].width = 20
    columnas['AA'].width = 18
    columnas['AB'].width = 18
    columnas['AC'].width = 10
    columnas['AD'].width = 25
    columnas['AE'].width = 18
    columnas['AF'].width = 8
    columnas['AG'].width = 18
    columnas['AH'].width = 8
    columnas['AI'].width = 18
    columnas['AJ'].width = 8
    columnas['AK'].width = 18
    columnas['AL'].width = 8
    columnas['AM'].width = 25
    columnas['AN'].width = 15
    columnas['AO'].width = 15
    columnas['AP'].width = 8
    columnas['AQ'].width = 15
    columnas['AR'].width = 20
    columnas['AS'].width = 15
    columnas['AT'].width = 30
    columnas['AU'].width = 15
    columnas['AV'].width = 20
    columnas['AW'].width = 30
    columnas['AX'].width = 30
    columnas['AY'].width = 15
    columnas['AZ'].width = 20
    columnas['BA'].width = 15
    columnas['BB'].width = 30
    columnas['BC'].width = 20
    columnas['BD'].width = 20
    for dato in personales:
       discapacitado =  dato.discapacitado
       if (discapacitado ==  True):
           discapacitado = 'SI'
       else:
           discapacitado = 'NO'
           
       estudias =  dato.estudias
       if (estudias ==  True):
           estudias = 'SI'
       else:
           estudias = 'NO' 
           
       comision =  dato.comision_servicio
       if (comision ==  True):
           comision = 'SI'
       else:
           comision = 'NO' 
           
       pnb =  dato.pnb
       if (pnb ==  True):
           pnb = 'SI'
       else:
           pnb = 'NO' 
           
       fasmij =  dato.fasmij
       if (fasmij ==  True):
           fasmij = 'SI'
       else:
           fasmij = 'NO' 
           
       ws.append([dato.estatus, dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.sexo.sexo, dato.fecha_nac, dato.edad, dato.telefono, dato.estado_civil.estado_civil, dato.conyugue, dato.cedula_conyugue, dato.tipo_sangre.tipo_sangre, discapacitado, dato.talla_camisa.talla_camisa, dato.talla_pantalon.talla_pantalon, dato.talla_zapato.talla_zapato, dato.direccion, dato.nro_cuenta, dato.email, dato.grado_instruccion.grado_instruccion, estudias, comision, pnb, dato.tipo_personal.tipo_personal, dato.cargo.cargo, dato.fecha_ingreso_911, dato.fecha_ingreso_apn, dato.contratos, dato.departamento.departamento, dato.nino_menor_12, dato.edades1, dato.hijos_13_18, dato.edades2, dato.nina_menor_12, dato.edades3, dato.hijos_discapacidad, dato.edades4, dato.motivo, dato.fecha_retiro, dato.sede.sede, fasmij, dato.parentezco1, dato.beneficiario1, dato.cedula1, dato.direccion1, dato.parentezco2, dato.beneficiario2, dato.cedula2, dato.direccion2, dato.parentezco3, dato.beneficiario3, dato.cedula3, dato.direccion3])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="datosPersonalRetirados.xlsx")

# DESCAGA PDF DE PERSONAL RETIRADO

def generar_pdf_retirados(request):
    # Consulta al modelo para obtener todos los datos
    datos = Personal.objects.filter(estatus='RETIRADO')

    # Crear el objeto HttpResponse con el tipo de contenido correcto
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="PersonalRetirado.pdf"'

    # Crear el objeto PDF
    pdf = SimpleDocTemplate(response, pagesize=letter, topMargin=20)
    elements = []
    
    
    # Agregar título centrado, en negrita y tamaño de letra 12
    title = "Reporte de Datos del Personal Retirado"
    title_style = getSampleStyleSheet()["Title"]
    title_style.alignment = 1  # 0=Left, 1=Center, 2=Right
    title_style.fontSize = 14
    title_text = Paragraph(title, title_style)
    elements.append(Spacer(1, 6))
    elements.append(title_text)
    
    # Agregar espacio en blanco entre el título y la tabla
    elements.append(Spacer(1, 2))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    # Agregar fecha debajo del título sin negrita y tamaño de letra 10
    now = datetime.datetime.now()
    fecha = "Fecha: %s / Hora: %s" % (now.strftime("%Y-%m-%d"), now.strftime("%H:%M"))
    fecha_style = getSampleStyleSheet()["Normal"]
    fecha_style.alignment = 1
    fecha_style.fontSize = 8
    fecha_text = Paragraph(fecha, fecha_style)
    elements.append(fecha_text)
    
    #Agregar espacio en blanco entre la fecha y la tabla
    elements.append(Spacer(1, 15))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    

    # Crear los datos para la tabla
    data = [['Nombres', 'Apellidos', 'Nacionalidad', 'Cédula', 'Teléfono', 'Cargo', 'Departamentos']]  # Encabezados de la tabla
    for dato in datos:

        discapacitado =  dato.discapacitado
        if (discapacitado ==  True):
           discapacitado = 'SI'
        else:
           discapacitado = 'NO'
           
        estudias =  dato.estudias
        if (estudias ==  True):
           estudias = 'SI'
        else:
           estudias = 'NO' 
           
        comision =  dato.comision_servicio
        if (comision ==  True):
           comision = 'SI'
        else:
           comision = 'NO' 
           
        pnb =  dato.pnb
        if (pnb ==  True):
           pnb = 'SI'
        else:
           pnb = 'NO' 
           
        fasmij =  dato.fasmij
        if (fasmij ==  True):
           fasmij = 'SI'
        else:
           fasmij = 'NO' 
           
        data.append([dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.telefono, dato.cargo.cargo])  # Agregar datos del modelo

    # Crear la tabla y aplicar estilos
    table = Table(data)
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    table.setStyle(style)

    # Agregar la tabla al PDF
    elements.append(table)

    # Construir el PDF y devolver la respuesta
    pdf.build(elements)
    return response









 #VISTA PERSONAL DE VACACIONES 
def personalVacaciones(request):
    personales = Personal.objects.filter(estatus='VACACIONES')
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "personalVacaciones.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades })


 #DESCAR EXCEL PERSONAL DE VACACIONES 
def descargar_excel_vacaciones(request):
    # Filtra los datos del modelo para generar el archivo Excel
    #personales = Personal.objects.all()
    personales = Personal.objects.filter(estatus='VACACIONES')

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:N1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Personal de Vacaciones del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    
    ws.append(["Estatus", "Nombres", "Apellidos", "Nacionalidad", "Cedula", "Genero", "Fecha Nac", "Edad", "Telefono", "Estado Civil", "Conyugue", "CI Conyugue", "Tipo Sangre", "Discapacitado", "Camisa", "Pantalon", "Zapatos", "Direccion", "Nro Cuenta", "Email", "Grado Instruccion", "Estudias", "Comision", "Pnb", "Tipo", "Cargo", "Fecha Ingreso 911", "Fecha Ingreso APN", "Contratos", "Departamentos", "Niños < 12", "Edad", "Hijos > 13", "Edad", "Niña < 12", "Edad", "Hijos Discapacidad", "Edad", "Motivo Retiro", "Fecha Retiro", "Sede", "Fasmij", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Fecha Creacion", "Fecha Actualizacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    columnas['F'].width = 15
    columnas['G'].width = 12
    columnas['H'].width = 5
    columnas['I'].width = 12
    columnas['J'].width = 12
    columnas['K'].width = 15
    columnas['L'].width = 10
    columnas['M'].width = 12
    columnas['N'].width = 12
    columnas['O'].width = 10
    columnas['P'].width = 10
    columnas['Q'].width = 10
    columnas['R'].width = 30
    columnas['S'].width = 25
    columnas['T'].width = 25
    columnas['U'].width = 17
    columnas['V'].width = 10
    columnas['W'].width = 10
    columnas['X'].width = 10
    columnas['Y'].width = 10
    columnas['Z'].width = 20
    columnas['AA'].width = 18
    columnas['AB'].width = 18
    columnas['AC'].width = 10
    columnas['AD'].width = 25
    columnas['AE'].width = 18
    columnas['AF'].width = 8
    columnas['AG'].width = 18
    columnas['AH'].width = 8
    columnas['AI'].width = 18
    columnas['AJ'].width = 8
    columnas['AK'].width = 18
    columnas['AL'].width = 8
    columnas['AM'].width = 25
    columnas['AN'].width = 15
    columnas['AO'].width = 15
    columnas['AP'].width = 8
    columnas['AQ'].width = 15
    columnas['AR'].width = 20
    columnas['AS'].width = 15
    columnas['AT'].width = 30
    columnas['AU'].width = 15
    columnas['AV'].width = 20
    columnas['AW'].width = 30
    columnas['AX'].width = 30
    columnas['AY'].width = 15
    columnas['AZ'].width = 20
    columnas['BA'].width = 15
    columnas['BB'].width = 30
    columnas['BC'].width = 20
    columnas['BD'].width = 20
    for dato in personales:
       discapacitado =  dato.discapacitado
       if (discapacitado ==  True):
           discapacitado = 'SI'
       else:
           discapacitado = 'NO'
           
       estudias =  dato.estudias
       if (estudias ==  True):
           estudias = 'SI'
       else:
           estudias = 'NO' 
           
       comision =  dato.comision_servicio
       if (comision ==  True):
           comision = 'SI'
       else:
           comision = 'NO' 
           
       pnb =  dato.pnb
       if (pnb ==  True):
           pnb = 'SI'
       else:
           pnb = 'NO' 
           
       fasmij =  dato.fasmij
       if (fasmij ==  True):
           fasmij = 'SI'
       else:
           fasmij = 'NO' 
           
       ws.append([dato.estatus, dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.sexo.sexo, dato.fecha_nac, dato.edad, dato.telefono, dato.estado_civil.estado_civil, dato.conyugue, dato.cedula_conyugue, dato.tipo_sangre.tipo_sangre, discapacitado, dato.talla_camisa.talla_camisa, dato.talla_pantalon.talla_pantalon, dato.talla_zapato.talla_zapato, dato.direccion, dato.nro_cuenta, dato.email, dato.grado_instruccion.grado_instruccion, estudias, comision, pnb, dato.tipo_personal.tipo_personal, dato.cargo.cargo, dato.fecha_ingreso_911, dato.fecha_ingreso_apn, dato.contratos, dato.departamento.departamento, dato.nino_menor_12, dato.edades1, dato.hijos_13_18, dato.edades2, dato.nina_menor_12, dato.edades3, dato.hijos_discapacidad, dato.edades4, dato.motivo, dato.fecha_retiro, dato.sede.sede, fasmij, dato.parentezco1, dato.beneficiario1, dato.cedula1, dato.direccion1, dato.parentezco2, dato.beneficiario2, dato.cedula2, dato.direccion2, dato.parentezco3, dato.beneficiario3, dato.cedula3, dato.direccion3])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="datosPersonalVacaciones.xlsx")

# DESCAGA PDF DE PERSONAL DE VACACIONES 

def generar_pdf_vacaciones(request):
    # Consulta al modelo para obtener todos los datos
    datos = Personal.objects.filter(estatus='VACACIONES')

    # Crear el objeto HttpResponse con el tipo de contenido correcto
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="PersonalDeVacaciones.pdf"'

    # Crear el objeto PDF
    pdf = SimpleDocTemplate(response, pagesize=letter, topMargin=20)
    elements = []
    
    
    # Agregar título centrado, en negrita y tamaño de letra 12
    title = "Reporte de Datos del Personal de Vacaciones"
    title_style = getSampleStyleSheet()["Title"]
    title_style.alignment = 1  # 0=Left, 1=Center, 2=Right
    title_style.fontSize = 14
    title_text = Paragraph(title, title_style)
    elements.append(Spacer(1, 6))
    elements.append(title_text)
    
    # Agregar espacio en blanco entre el título y la tabla
    elements.append(Spacer(1, 2))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    # Agregar fecha debajo del título sin negrita y tamaño de letra 10
    now = datetime.datetime.now()
    fecha = "Fecha: %s / Hora: %s" % (now.strftime("%Y-%m-%d"), now.strftime("%H:%M"))
    fecha_style = getSampleStyleSheet()["Normal"]
    fecha_style.alignment = 1
    fecha_style.fontSize = 8
    fecha_text = Paragraph(fecha, fecha_style)
    elements.append(fecha_text)
    
    #Agregar espacio en blanco entre la fecha y la tabla
    elements.append(Spacer(1, 15))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    

    # Crear los datos para la tabla
    data = [['Nombres', 'Apellidos', 'Nacionalidad', 'Cédula', 'Teléfono', 'Cargo', 'Departamentos']]  # Encabezados de la tabla
    for dato in datos:

        discapacitado =  dato.discapacitado
        if (discapacitado ==  True):
           discapacitado = 'SI'
        else:
           discapacitado = 'NO'
           
        estudias =  dato.estudias
        if (estudias ==  True):
           estudias = 'SI'
        else:
           estudias = 'NO' 
           
        comision =  dato.comision_servicio
        if (comision ==  True):
           comision = 'SI'
        else:
           comision = 'NO' 
           
        pnb =  dato.pnb
        if (pnb ==  True):
           pnb = 'SI'
        else:
           pnb = 'NO' 
           
        fasmij =  dato.fasmij
        if (fasmij ==  True):
           fasmij = 'SI'
        else:
           fasmij = 'NO' 
           
        data.append([dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.telefono, dato.cargo.cargo])  # Agregar datos del modelo

    # Crear la tabla y aplicar estilos
    table = Table(data)
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    table.setStyle(style)

    # Agregar la tabla al PDF
    elements.append(table)

    # Construir el PDF y devolver la respuesta
    pdf.build(elements)
    return response




    
    
def bienes(request, accion, departamentofilter):

    bienes = Bienes.objects.all()
    
    #   CONSULTA POR TODOS 
    if accion == 'todos' and departamentofilter != 0:
        bienes = Bienes.objects.filter(departamento=departamentofilter)
        form = FormBienes()
    else:
        form = FormBienes()
        
    #   AGREGAR BIENES
    if accion == 'agregar':
        if request.method =='POST':
            form = FormBienes(request.POST)
            if form.is_valid():
                form.save()
    
    #   CONSULTAR POR NOMBRE DEL BIEN
    if accion == 'consultar':
        if 'nombre' in request.GET:
            nombre = request.GET['nombre']
            bienes = Bienes.objects.filter(nombre=nombre)
            form = FormBienes()
            
    #   EDICION DE BIENES
    if accion == 'editar':
        if request.method == 'POST':
            form = FormBienes(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                nombre_m = form.cleaned_data['nombre']
                cantidad_m = form.cleaned_data['cantidad']
                departamento_m = form.cleaned_data['departamento']
                sede_m = form.cleaned_data['sede']
                
                try:
                    bien = Bienes.objects.get(id=id_m)
                except Bienes.DoesNotExist:
                    raise ValueError("Bien no existe.")
                    
                bien.estatus = estatus_m
                bien.sede = sede_m
                bien.departamento = departamento_m
                bien.nombre = nombre_m
                bien.cantidad = cantidad_m
                bien.save()
        else:
            form = FormBienes()
            
    sedes = sede()
    departamentos = departamento()
    context = {
        "bienes": bienes,
        "sedes": sedes,
        "departamentos": departamentos,
        "form": form,
    }
    return render(request, "bienes.html", context)


def obtener_datos_bienes_edicion(request, bienes_id):
    try:
        bienes = Bienes.objects.get(id=bienes_id)
        data = {
            'id': bienes.id,
            'estatus': bienes.estatus,
            'nombre': bienes.nombre,
            'cantidad': bienes.cantidad,
            'departamento': bienes.departamento.id,
            'sede': bienes.sede.id,
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    
    
def descargar_excel_bienes(request):
    # Filtra los datos del modelo para generar el archivo Excel
    bienes = Bienes.objects.all()

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:E1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Bienes Nacionales del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    
    ws.append(["Estatus", "Sede", "Departamento", "Nombre", "Cantidad"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    
    for dato in bienes:
       
       ws.append([dato.estatus, dato.sede.sede, dato.departamento.departamento, dato.nombre, dato.cantidad])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="bienesNacionales.xlsx")


    #DESCAR PDF BIENES NACIONALES 
def generar_pdf_bienes(request):
    # Consulta al modelo para obtener todos los datos
    datos = Bienes.objects.all()

    # Crear el objeto HttpResponse con el tipo de contenido correcto
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="BienesNacionales.pdf"'

    # Crear el objeto PDF
    pdf = SimpleDocTemplate(response, pagesize=letter, topMargin=20)
    elements = []
    
    
    # Agregar título centrado, en negrita y tamaño de letra 12
    title = "Reporte de Bienes Nacionales"
    title_style = getSampleStyleSheet()["Title"]
    title_style.alignment = 1  # 0=Left, 1=Center, 2=Right
    title_style.fontSize = 14
    title_text = Paragraph(title, title_style)
    elements.append(Spacer(1, 6))
    elements.append(title_text)
    
    # Agregar espacio en blanco entre el título y la tabla
    elements.append(Spacer(1, 2))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    # Agregar fecha debajo del título sin negrita y tamaño de letra 10
    now = datetime.datetime.now()
    fecha = "Fecha: %s / Hora: %s" % (now.strftime("%Y-%m-%d"), now.strftime("%H:%M"))
    fecha_style = getSampleStyleSheet()["Normal"]
    fecha_style.alignment = 1
    fecha_style.fontSize = 8
    fecha_text = Paragraph(fecha, fecha_style)
    elements.append(fecha_text)
    
    #Agregar espacio en blanco entre la fecha y la tabla
    elements.append(Spacer(1, 15))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    

    # Crear los datos para la tabla
    data = [[ 'Estatus', 'Sede', 'Departamento', 'Nombre', 'Cantidad', 'Creado']]  # Encabezados de la tabla
    for dato in datos:          
        data.append([dato.estatus, dato.sede.sede, dato.departamento.departamento, dato.nombre, dato.cantidad, dato.created])  # Agregar datos del modelo

    # Crear la tabla y aplicar estilos
    table = Table(data)
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    table.setStyle(style)

    # Agregar la tabla al PDF
    elements.append(table)

    # Construir el PDF y devolver la respuesta
    pdf.build(elements)
    return response



def sueldos(request):

    sueldos = Sueldos.objects.all()
    tasas = Tasa.objects.first()
    
    resultadoscesta = []
    resultadosbono = []
    
    for sueldo in sueldos:
        resultadoc = float(sueldo.monto_t.monto) * float(tasas.monto)
        resultadoscesta.append(resultadoc)
        resultadob = float(sueldo.monto_b.monto) * float(tasas.monto)
        resultadosbono.append(resultadob)
        
    datos_combinados = zip(sueldos, resultadoscesta, resultadosbono)

        
    context = {
        "sueldos": sueldos,
        "datos_combinados": datos_combinados
    }
    return render(request, "sueldos.html", context)

def obtener_datos_sueldos_edicion(request, sueldo_id):
    try:
        sueldo=Sueldos.objects.get(id=sueldo_id)
        data = {
            'id': sueldo.id,
            'estatus': sueldo.estatus,
            'sueldo_base': sueldo.sueldo_base,
            'prima_profesionalismo': sueldo.prima_profesionalismo,
            'p_discapacidad': sueldo.p_discapacidad,
            'p_hijos_menor_12' : sueldo.p_hijos_menor_12,
            'p_hijas_menor_12' : sueldo.p_hijas_menor_12,
            'p_hijos_12_18' : sueldo.p_hijos_12_18,
            'p_hijos_discapacidad' : sueldo.p_hijos_discapacidad,
            'n_fasmij' : sueldo.n_fasmij,
            'cesta_t' : sueldo.cesta_t,
            'monto_t' : sueldo.monto_t.monto,
            'b_guerra' : sueldo.b_guerra,
            'monto_b' : sueldo.monto_b.monto, 
            'personal' : sueldo.personal.nombres, 
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  
    
    #DESCARGAR PDF DE SUELDOS 

def generar_pdf_sueldos(request):
    # Consulta al modelo para obtener todos los datos
    datos = Sueldos.objects.all()
    tasa = Tasa.objects.first()
    
    tasabcv = tasa.monto

    # Crear el objeto HttpResponse con el tipo de contenido correcto
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Salarios.pdf"'

    # Crear el objeto PDF
    pdf = SimpleDocTemplate(response, pagesize=letter, topMargin=20)
    elements = []
    
    
    # Agregar título centrado, en negrita y tamaño de letra 12
    title = "Reporte de Salarios"
    title_style = getSampleStyleSheet()["Title"]
    title_style.alignment = 1  # 0=Left, 1=Center, 2=Right
    title_style.fontSize = 14
    title_text = Paragraph(title, title_style)
    elements.append(Spacer(1, 6))
    elements.append(title_text)
    
    # Agregar espacio en blanco entre el título y la tabla
    elements.append(Spacer(1, 2))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    # Agregar fecha debajo del título sin negrita y tamaño de letra 10
    now = datetime.datetime.now()
    fecha = "Fecha: %s / Hora: %s" % (now.strftime("%Y-%m-%d"), now.strftime("%H:%M"))
    fecha_style = getSampleStyleSheet()["Normal"]
    fecha_style.alignment = 1
    fecha_style.fontSize = 8
    fecha_text = Paragraph(fecha, fecha_style)
    elements.append(fecha_text)
    
    #Agregar espacio en blanco entre la fecha y la tabla
    elements.append(Spacer(1, 15))  # Puedes ajustar el tamaño del espacio según tus necesidades
    
    

    # Crear los datos para la tabla
    data = [[ 'Nombres', 'Apellidos', 'Cédula', 'Cargo', 'Sueldo Base', 'Ticket', 'Bono de Guerra']]  # Encabezados de la tabla
    for dato in datos:         
        ticket = dato.monto_t.monto * tasabcv
        ticket = "{:.2f} Bs.".format(ticket)
        bono = dato.monto_b.monto * tasabcv
        bono = "{:.2f} Bs.".format(bono)
        data.append([ dato.personal.nombres, dato.personal.apellidos, dato.personal.cedula, dato.personal.cargo.cargo, dato.sueldo_base, ticket, bono])  # Agregar datos del modelo
    

    # Crear la tabla y aplicar estilos
    table = Table(data)
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    table.setStyle(style)

    # Agregar la tabla al PDF
    elements.append(table)

    # Construir el PDF y devolver la respuesta
    pdf.build(elements)
    return response

#DESCARGA EXCEL DE SUELDOS 
def descargar_excel_sueldos(request):
    # Filtra los datos del modelo para generar el archivo Excel
    sueldos = Sueldos.objects.all()
    tasa = Tasa.objects.first()
    
    tasabcv = tasa.monto

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:E1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Salarios del personal'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    
    # Obtener la columna G
    column_g = ws['G']

    # Iterar sobre cada celda de la columna G
    for cell in column_g:
        cell.alignment = Alignment(horizontal='right')
    
    ws.append(["Nombres", "Apellidos", "Cédula", "Cargo", "Sueldo Base", "Ticket", "Bono de Guerra"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 15
    columnas['C'].width = 15
    columnas['D'].width = 20
    columnas['E'].width = 15
    columnas['F'].width = 15
    columnas['G'].width = 15
    
    for dato in sueldos:
        ticket = dato.monto_t.monto * tasabcv
        ticket = "{:.2f} Bs.".format(ticket)
        bono = dato.monto_b.monto * tasabcv
        bono = "{:.2f} Bs.".format(bono)
       
        ws.append([dato.personal.nombres, dato.personal.apellidos, dato.personal.cedula, dato.personal.cargo.cargo, dato.sueldo_base, ticket, bono])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="Salarios.xlsx")



def mantBonoGuerra(request, accion):

    bonoguerra=BonoGuerra.objects.first()
    form = FormBonoGuerra()
    
    if accion == 'todos':
        form = FormBonoGuerra()


    #EDITAR BONO   
    if accion == 'editar':
        if request.method == 'POST':
            form = FormBonoGuerra(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                #descripcion_m = form.cleaned_data['descripcion']
                monto_m = form.cleaned_data['monto']
                
                try:
                    bonoG = BonoGuerra.objects.get(id=id_m)
                except BonoGuerra.DoesNotExist:
                    raise ValueError("Bono  no existe.")
                    
                bonoG.estatus = estatus_m
                #bonoG.descripcion = descripcion_m
                bonoG.monto = monto_m
                bonoG.save()
                return redirect('/rrhh/bonoguerra/todos')
        else:
             
             form = FormBonoGuerra()
            
    
    context = {
        "bonoguerra": bonoguerra,
        "form": form
    }

    return render(request, "mantenimientos/bonoGuerra.html", context)



def obtener_datos_bono_edicion(request, bono_id):
    try:
        bonoguerra=BonoGuerra.objects.get(id=bono_id)
        data = {
            'id': bonoguerra.id,
            'estatus': bonoguerra.estatus,
            'descripcion': bonoguerra.descripcion,
            'monto': bonoguerra.monto
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  



def mantCestaTicket(request, accion):

    cestaticket=CestaTicket.objects.first()
    form = FormCestaTicket()
    
    if accion == 'todos':
        form = FormCestaTicket()

    #EDITAR CESTATICKET
    if accion == 'editar':
        if request.method == 'POST':
            form = FormCestaTicket(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                #descripcion_m = form.cleaned_data['descripcion']
                monto_m = form.cleaned_data['monto']
                
                try:
                    cestaT = CestaTicket.objects.get(id=id_m)
                except CestaTicket.DoesNotExist:
                    raise ValueError("Cesta  no existe.")
                    
                cestaT.estatus = estatus_m
                cestaT.monto = monto_m
                cestaT.save()
                return redirect('/rrhh/cestaticket/todos')
        else:
             
             form = FormCestaTicket()
            
    
    context = {
        "cestaticket": cestaticket,
        "form": form
    }

    return render(request, "mantenimientos/cestaTicket.html", context)



def obtener_datos_cestaticket_edicion(request, cesta_id):
    try:
        cestaticket=CestaTicket.objects.get(id=cesta_id)
        data = {
            'id': cestaticket.id,
            'estatus': cestaticket.estatus,
            'descripcion': cestaticket.descripcion,
            'monto': cestaticket.monto
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  
    

    
def mantSueldoMinimo(request, accion):

    sueldominimo=SueldoMinimo.objects.first()
    form = FormSueldoMinimo()
    
    if accion == 'todos':
        form = FormSueldoMinimo()

    #EDITAR CESTATICKET
    if accion == 'editar':
        if request.method == 'POST':
            form = FormSueldoMinimo(request.POST)
            if form.is_valid():
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                #descripcion_m = form.cleaned_data['descripcion']
                monto_m = form.cleaned_data['monto']
                
                try:
                    sueldoM = SueldoMinimo.objects.get(id=id_m)
                except SueldoMinimo.DoesNotExist:
                    raise ValueError("Cesta  no existe.")
                    
                sueldoM.estatus = estatus_m
                sueldoM.monto = monto_m
                sueldoM.save()
                return redirect('/rrhh/sueldominimo/todos')
        else:
             
             form = FormCestaTicket()
            
    
    context = {
        "sueldominimo": sueldominimo,
        "form": form
    }

    return render(request, "mantenimientos/sueldoMinimo.html", context)


def obtener_datos_sueldominimo_edicion(request, sueldo_id):
    try:
        sueldominimo=SueldoMinimo.objects.get(id=sueldo_id)
        data = {
            'id': sueldominimo.id,
            'estatus': sueldominimo.estatus,
            'descripcion': sueldominimo.descripcion,
            'monto': sueldominimo.monto
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except Personal.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)  
    

def mantPrimaProfesional(request, accion):
    primasProfesionales=PrimaProfesionalismo.objects.all()
    grado_instruccion=Grado.objects.all()
    
    form = FormPrima()
    novalido = "Formulario no valido"
    valido = "Formulario valido"
    
    if accion == 'todos':
        form = FormPrima()
        
    #   AGREGAR PRIMA
    if accion == 'agregar':
        if request.method =='POST':
            formAdd = FormPrimaAdd(request.POST)
            print(formAdd)
            if formAdd.is_valid():
                print (valido)
                formAdd.save()
                return redirect('/rrhh/primaprofesional/todos')
            else:
                print(novalido)
                
    #EDITAR PRIMA
    if accion == 'editar':
        if request.method == 'POST':
            form = FormPrima(request.POST)
            if form.is_valid():
                #print (valido)
                id_m =  request.POST["id"]
                estatus_m = form.cleaned_data['estatus']
                #titulo = form.cleaned_data['titulo']
                porcentaje_m = form.cleaned_data['porcentaje']
                
                try:
                    primaP = PrimaProfesionalismo.objects.get(id=id_m)
                except PrimaProfesionalismo.DoesNotExist:
                    raise ValueError("Prima  no existe.")
                    
                primaP.estatus = estatus_m
                primaP.porcentaje = porcentaje_m
                #print(porcentaje_m)
                primaP.save()
                return redirect('/rrhh/primaprofesional/todos')
            else:
                print(novalido)
        else:
             
             form = FormPrima()
            
    
    context = {
        "primasProfesionales": primasProfesionales,
        "grado_instruccion":grado_instruccion,
        "form": form
    }

    return render(request, "mantenimientos/prima_profesional.html", context)



def obtener_datos_prima_edicion(request, prima_id):
    try:
        prima = PrimaProfesionalismo.objects.get(id=prima_id)
        data = {
            'id': prima.id,
            'estatus': prima.estatus,
            'titulo': prima.titulo.id,
            'porcentaje': prima.porcentaje,
            
            # Agrega más campos según sea necesario
        }
        return JsonResponse(data)
    except PrimaProfesionalismo.DoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    